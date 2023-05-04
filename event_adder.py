import openpyxl
from raiseredge import RaiserEdge 

'''
Creating Event Script
Make sure the spreadsheet has the rows following this order with the same name: "Event_name", "Event_Date", "Event_Time", "ID"
Make sure to add your access_key and oauth key in Line 25
'''

# Add Event Path
path = ""
wb = openpyxl.load_workbook(path)
sheet = wb.active
max_col = sheet.max_column

for i in range(1, max_col + 1):
    try:
        name_list = ['Event_name', 'Event_Date', 'Event_Time', 'ID']
        assert sheet.cell(row = 1, column = i).value == name_list[i - 1]
    except:
        raise Exception("Only four rows in this order: Event_name, Event_Date, Event_Time and ID")

max_row = sheet.max_row

re = RaiserEdge(access_key='', oauth='')

for i in range(2, max_row + 1):
    if sheet.cell(row = i, column = 1) != 'Event_name':
        json_data = {
                    "name": f"{sheet.cell(row = i, column = 1).value}",
                    "start_date": str(sheet.cell(row = i, column = 2).value)[:10],
                    "category": {
                        "id": "",
                        "name": ""
                    },
                    "start_time": str(sheet.cell(row = i, column = 3).value)[:5],
                    "lookup_id": str(sheet.cell(row = i, column = 4).value),
                    "description": None,
                    "end_date": None,
                    "inactive": False,
                    "capacity": None,
                    "goal": None,
                    "campaign_id": None,
                    "fund_id": None
                }

        print("Imported: " + sheet.cell(row = i, column = 1).value)
        print(re.create_event(**json_data))





