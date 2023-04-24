import openpyxl
from raiseredge import RaiserEdge 

access = ''
auth = ''
# re = RaiserEdge(access_key=access, oauth=auth)


def is_net_id(s):
    
    if(len(s) != 6 or not s[:2].isalpha() or not s[2:].isdigit()):
        return False
    return True

def get_event_id_from_lookup_id(id:str):
    event_list_params = {
        'lookup_id': id
    }
    
    x = re.get_event_list(**event_list_params)
    return x["value"][0]['id']

def find_constituent_id(first_name: str, last_name: str, email: str, net_id: str):
    if email == None:
        return None

    search_email = {
        'search_text': email,
        'search_field': 'email_address'
    }
    constituent_email = re.search_constituent(**search_email)

    if constituent_email['count'] > 0:
        for i in range(int(constituent_email['count'])):
    
            if first_name.capitalize in str(constituent_email['value'][i]['name']).capitalize and last_name.capitalize in str((constituent_email['value'][i]['name'])).capitalize:
                try:
                    return constituent_email['value'][i]['lookup_id']
                except:
                    print("Bruh")

    if net_id == None:
        return None

    search_email = {
        'search_text': net_id
    }
    constituent_netid = re.search_constituent(**search_email)
    
    if constituent_netid['count'] > 0:
        return constituent_netid['value'][0]['lookup_id']

    return None
if __name__ == "__main__":
    

    path = "C:\\Users\\STSC\\Documents\\Programming\\Python\\Raiser's Edge NXT API Wrapper\\MasterEventList 20-21 (3).xlsx"
    
    
    

    wb = openpyxl.load_workbook(path)
    # sheets = wb.sheetnames
    # print(sheets)

    sheet = wb["PGN 6.24"]
    for i in range(8, 34):
        net_id = str(sheet.cell(row = i, column = 7).value)
        if not is_net_id(net_id):
            net_id = None
        print(str(sheet.cell(row = i, column = 2).value), str(sheet.cell(row = i, column = 3).value), sheet.cell(row = i, column = 4).value, net_id)

'''
    for sheet in sheets:
        temp_sheet = wb.get_sheet_by_name(sheet)
        print(sheet, str((temp_sheet.cell(row = 7, column = 1).value == 'Attendee Status')))

'''


